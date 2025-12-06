"""
document_processor.py FINAL (2025)
-----------------------------------------------------
Orden de lectura:
 1. Simple moderno (docx, xlsx, xls, pdf, txt, csv)
 2. Legacy Reader (ppt antiguos, doc antiguos, xls antiguos)
 3. OCR fallback
-----------------------------------------------------

Incluye mejoras del modo IA de la app Android:
 - Limpieza profunda de texto
 - División en bloques igual que la app
 - Cálculo dinámico de preguntas
 - Recorte por tipo (BASICO, MODERADO, COMPLETO)
"""

from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader
import os
import re
import xlrd

# Legacy readers
import legacy_reader


class DocumentProcessor:

    # ======================================================
    # 🔹 1. LECTOR SIMPLE (modernos)
    # ======================================================
    def leer_simple(self, path, extension):
        try:
            # DOCX
            if extension == "docx":
                doc = Document(path)
                texto = "\n".join([p.text for p in doc.paragraphs])
                return {"texto": texto, "method": "docx", "warnings": []}

            # XLSX
            elif extension == "xlsx":
                libro = load_workbook(filename=path, data_only=True)
                texto = ""
                for hoja in libro.sheetnames:
                    ws = libro[hoja]
                    for row in ws.iter_rows():
                        texto += " ".join([str(c.value) if c.value else "" for c in row]) + "\n"
                return {"texto": texto, "method": "xlsx", "warnings": []}

            # XLS moderno
            elif extension == "xls":
                texto = ""
                libro = xlrd.open_workbook(path)
                for hoja in libro.sheets():
                    for row_idx in range(hoja.nrows):
                        fila = hoja.row_values(row_idx)
                        texto += " ".join([str(v) for v in fila]) + "\n"
                return {"texto": texto, "method": "xls", "warnings": []}

            # TXT
            elif extension == "txt":
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    return {"texto": f.read(), "method": "txt", "warnings": []}

            # CSV
            elif extension == "csv":
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    return {"texto": f.read(), "method": "csv", "warnings": []}

            return {"texto": "", "method": "simple_unknown", "warnings": []}

        except Exception as e:
            return {"texto": "", "method": "simple_error", "warnings": [str(e)]}

    # ======================================================
    # 🔹 2. LECTOR PDF
    # ======================================================
    def leer_pdf(self, path):
        try:
            reader = PdfReader(path)
            texto = ""
            for page in reader.pages:
                texto += page.extract_text() or ""

            return {"texto": texto, "method": "pdf", "warnings": []}
        except Exception as e:
            return {"texto": "", "method": "pdf_error", "warnings": [str(e)]}

    # ======================================================
    # 🔹 3. LECTOR LEGACY
    # ======================================================
    def leer_legacy(self, path, extension):
        try:
            if extension == "ppt":
                texto = legacy_reader.leer_ppt_antiguo(path)
                return {"texto": texto, "method": "legacy_ppt", "warnings": []}

            if extension == "doc":
                texto = legacy_reader.leer_doc_antiguo(path)
                return {"texto": texto, "method": "legacy_doc", "warnings": []}

            if extension == "xls":
                texto = legacy_reader.leer_xls_antiguo(path)
                return {"texto": texto, "method": "legacy_xls", "warnings": []}

        except Exception as e:
            return {"texto": "", "method": "legacy_error", "warnings": [str(e)]}

        return {"texto": "", "method": "legacy_none", "warnings": []}

    # ======================================================
    # 🔹 4. OCR UNIVERSAL
    # ======================================================
    def ocr(self, path):
        try:
            texto = legacy_reader.ocr_fallback(path)
            return {"texto": texto, "method": "ocr", "warnings": []}
        except Exception as e:
            return {"texto": "", "method": "ocr_error", "warnings": [str(e)]}

    # ======================================================
    # 🔹 5. LIMPIEZA DE TEXTO (nuevo, igual a la app)
    # ======================================================
    def limpiar_texto(self, texto: str):
        texto = texto.replace("\u200b", "").replace("\ufeff", "")
        texto = re.sub(r"(OCR|PAGINA_\d+|ERROR|FAILED|SCAN)", "", texto, flags=re.I)
        texto = re.sub(r"\s+", " ", texto)
        return texto.strip()

    # ======================================================
    # 🔹 6. DIVISIÓN EN BLOQUES (igual a la app Android)
    # ======================================================
    def dividir_en_bloques(self, texto, min_size=1200, max_size=2000):

        lineas = texto.split("\n")
        bloques = []
        actual = ""

        for linea in lineas:
            if len(actual) + len(linea) > max_size:
                bloques.append(actual.strip())
                actual = linea
            else:
                actual += " " + linea

        if actual.strip():
            bloques.append(actual.strip())

        # Unir bloque final pequeño (<300 chars)
        if len(bloques) >= 2 and len(bloques[-1]) < 300:
            bloques[-2] += " " + bloques[-1]
            bloques.pop()

        return bloques

    # ======================================================
    # 🔹 7. CÁLCULO DE PREGUNTAS POR BLOQUE
    # ======================================================
    def calcular_preguntas_por_bloque(self, total_bloques):
        try:
            base = int(200 / max(1, total_bloques))
            return max(8, min(25, base))
        except:
            return 8

    # ======================================================
    # 🔹 8. RECORTE POR TIPO DE TEST (opcional)
    # ======================================================
    def recortar_por_tipo(self, texto, tipo):
        """
        Tipos:
        - basico → 1500 caracteres
        - moderado → 4000 caracteres
        - completo → 10000 caracteres
        """

        limites = {
            "basico": 1500,
            "moderado": 4000,
            "completo": 10000
        }

        limite = limites.get(tipo.lower(), 10000)
        return texto[:limite]

    # ======================================================
    # 🔹 9. ORDEN PRINCIPAL
    # ======================================================
    def procesar_archivo(self, path, extension):
        extension = extension.lower()

        # 1️⃣ PDF directo
        if extension == "pdf":
            r_pdf = self.leer_pdf(path)
            if r_pdf["texto"].strip():
                return r_pdf

        # 2️⃣ Simple reader
        r_simple = self.leer_simple(path, extension)
        if r_simple["texto"].strip():
            return r_simple

        # 3️⃣ Legacy reader
        r_legacy = self.leer_legacy(path, extension)
        if r_legacy["texto"].strip():
            return r_legacy

        # 4️⃣ OCR fallback
        r_ocr = self.ocr(path)
        if r_ocr["texto"].strip():
            return r_ocr

        # Nada funcionó
        return {"texto": "", "method": "none", "warnings": ["No se pudo leer el documento"]}

